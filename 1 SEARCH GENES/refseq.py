import random
import time
from pathlib import Path

import ssl
import certifi

from Bio import Entrez, SeqIO
from Bio.Seq import Seq
from Bio.SeqRecord import SeqRecord

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# --- SSL (Windows-friendly) ---
ssl_context = ssl.create_default_context(cafile=certifi.where())
ssl._create_default_https_context = lambda: ssl_context

# --- NCBI etiquette ---
# We'll adapt this at runtime depending on whether an API key is provided.
NCBI_REQUEST_DELAY_SECONDS = 1 / 3

# --- Retry tuning (covers both network calls and mid-stream parse EOF) ---
ENTREZ_MAX_RETRIES = 10
ENTREZ_BACKOFF_BASE_SECONDS = 1.8
ENTREZ_JITTER_SECONDS = 0.6

# --- Batch sizes ---
ASSEMBLY_RETMAX = 50
NUCCORE_LINK_RETMAX = 200
GB_FETCH_BATCH = 5  # keep small: gbwithparts can be large and more likely to EOF


def prompt_required(prompt_text: str) -> str:
    while True:
        v = input(prompt_text).strip()
        if v:
            return v
        print("This field is required.")


def prompt_optional_int(prompt_text: str) -> int | None:
    v = input(prompt_text).strip()
    if not v:
        return None
    try:
        return int(v)
    except ValueError:
        print("Invalid number. Ignoring.")
        return None


def prompt_optional_str(prompt_text: str) -> str | None:
    v = input(prompt_text).strip()
    return v or None


def normalize_gene_terms(raw_genes: str) -> list[str]:
    return [g.strip() for g in raw_genes.split(",") if g.strip()]


def is_transient_entrez_error(exc: Exception) -> bool:
    msg = str(exc).lower()
    transient_markers = [
        "read failed",
        "eof",
        "closed connection",
        "unexpectedly closed",
        "connection reset",
        "connection aborted",
        "timed out",
        "timeout",
        "temporary failure",
        "temporarily unavailable",
        "try again",
        "server error",
        "bad gateway",
        "service unavailable",
        "gateway timeout",
        "http error 5",
        "502",
        "503",
        "504",
    ]
    return any(m in msg for m in transient_markers)


def backoff_sleep(attempt: int) -> None:
    wait = (ENTREZ_BACKOFF_BASE_SECONDS ** (attempt - 1)) + random.uniform(0, ENTREZ_JITTER_SECONDS)
    time.sleep(wait)


def build_assembly_query(organism: str | None, taxid: str | None) -> str:
    terms = ["(latest_refseq[filter])"]
    if taxid:
        terms.append(f"(txid{taxid}[Organism:exp])")
    elif organism:
        terms.append(f"({organism}[Organism])")
    return " AND ".join(terms)


def esearch_ids(db: str, term: str, retmax: int) -> list[str]:
    last_exc: Exception | None = None
    for attempt in range(1, ENTREZ_MAX_RETRIES + 1):
        try:
            time.sleep(NCBI_REQUEST_DELAY_SECONDS)
            h = Entrez.esearch(db=db, term=term, retmax=retmax)
            r = Entrez.read(h)
            h.close()
            return r.get("IdList", [])
        except Exception as exc:
            last_exc = exc
            if not is_transient_entrez_error(exc):
                raise
            print(f"\nWarning: transient error during esearch({db}) attempt {attempt}/{ENTREZ_MAX_RETRIES}: {exc}")
            print("Retrying...")
            backoff_sleep(attempt)
    raise last_exc  # type: ignore[misc]


def elink_assembly_to_nuccore_ids(assembly_id: str) -> list[str]:
    last_exc: Exception | None = None
    for attempt in range(1, ENTREZ_MAX_RETRIES + 1):
        try:
            time.sleep(NCBI_REQUEST_DELAY_SECONDS)
            h = Entrez.elink(dbfrom="assembly", db="nuccore", id=assembly_id, linkname="assembly_nuccore_refseq")
            r = Entrez.read(h)
            h.close()

            ids: list[str] = []
            for linkset in r:
                for linksetdb in linkset.get("LinkSetDb", []):
                    for link in linksetdb.get("Link", []):
                        ids.append(link["Id"])
            return ids[:NUCCORE_LINK_RETMAX]
        except Exception as exc:
            last_exc = exc
            if not is_transient_entrez_error(exc):
                raise
            print(f"\nWarning: transient error during elink attempt {attempt}/{ENTREZ_MAX_RETRIES}: {exc}")
            print("Retrying...")
            backoff_sleep(attempt)
    raise last_exc  # type: ignore[misc]


def fetch_genbank_batch_records_with_retry(nuccore_ids_batch: list[str]):
    """
    Critical: retry the *whole* fetch+parse, because EOF can occur mid-parse.
    """
    last_exc: Exception | None = None
    for attempt in range(1, ENTREZ_MAX_RETRIES + 1):
        h = None
        try:
            time.sleep(NCBI_REQUEST_DELAY_SECONDS)
            h = Entrez.efetch(
                db="nuccore",
                id=",".join(nuccore_ids_batch),
                rettype="gbwithparts",
                retmode="text",
            )

            # Materialize immediately so parsing errors are caught here and retried.
            records = list(SeqIO.parse(h, "genbank"))
            h.close()
            return records

        except Exception as exc:
            last_exc = exc
            try:
                if h is not None:
                    h.close()
            except Exception:
                pass

            if not is_transient_entrez_error(exc):
                raise

            print(
                f"\nWarning: transient error during efetch+parse gbwithparts "
                f"(attempt {attempt}/{ENTREZ_MAX_RETRIES}): {exc}"
            )
            print("Retrying this batch...")
            backoff_sleep(attempt)

    raise last_exc  # type: ignore[misc]


def feature_matches_gene(feature, gene_terms_lower: set[str]) -> tuple[bool, str]:
    quals = feature.qualifiers

    def any_match(values):
        for v in values:
            v_low = v.lower()
            for g in gene_terms_lower:
                if g == v_low or g in v_low:
                    return True
        return False

    if "gene" in quals and any_match(quals["gene"]):
        return True, "gene"
    if "gene_synonym" in quals and any_match(quals["gene_synonym"]):
        return True, "gene_synonym"
    if "product" in quals and any_match(quals["product"]):
        return True, "product"
    if "note" in quals and any_match(quals["note"]):
        return True, "note"
    return False, ""


def extract_cds_nucleotide(seq_record: SeqRecord, cds_feature):
    try:
        return cds_feature.extract(seq_record.seq)
    except Exception:
        return None


def extract_protein_from_feature(cds_feature) -> str | None:
    tr = cds_feature.qualifiers.get("translation")
    if tr and len(tr) > 0:
        return tr[0]
    return None


def make_fasta_header(genome_accession: str, organism: str, gene_label, product, protein_id, locus_tag) -> str:
    parts = [genome_accession]
    if gene_label:
        parts.append(f"gene={gene_label}")
    if locus_tag:
        parts.append(f"locus_tag={locus_tag}")
    if protein_id:
        parts.append(f"protein_id={protein_id}")
    if product:
        parts.append(f"product={product}")
    if organism:
        parts.append(f"organism={organism}")
    return " | ".join(parts)


def write_excel(rows: list[dict], out_xlsx: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "QC"

    headers = [
        "organism",
        "genome_accession",
        "code",
        "record_description",
        "matched_by",
        "gene",
        "locus_tag",
        "product",
        "protein_id",
        "cds_location",
        "strand",
        "length",
    ]

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="2E4057")
    centered = Alignment(horizontal="center", vertical="center")

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered

    for r_i, row in enumerate(rows, start=2):
        for c_i, h in enumerate(headers, start=1):
            ws.cell(row=r_i, column=c_i, value=row.get(h, ""))

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 35
    ws.column_dimensions["I"].width = 18
    ws.column_dimensions["J"].width = 30
    ws.column_dimensions["K"].width = 8
    ws.column_dimensions["L"].width = 10

    wb.save(out_xlsx)


def main():
    global NCBI_REQUEST_DELAY_SECONDS

    print("=== Extract Gene by Annotation ===")
    print("This extracts CDS by gene/product annotation from RefSeq genome records in NCBI.\n")

    Entrez.email = prompt_required("Email for NCBI Entrez (required by NCBI): ")
    Entrez.tool = "refseq_gene_extractor"

    api_key = prompt_optional_str("NCBI API key (optional, recommended): ")
    if api_key:
        Entrez.api_key = api_key
        # safer than 0.1 for heavy gbwithparts; still fast enough
        NCBI_REQUEST_DELAY_SECONDS = 0.15
    else:
        NCBI_REQUEST_DELAY_SECONDS = 1 / 3

    search_for = input("Search for (nucleotide | protein) sequence: ").strip().lower() or "nucleotide"
    if search_for not in {"nucleotide", "protein"}:
        search_for = "nucleotide"
    extract_mode = "protein" if search_for == "protein" else "cds_nt"

    if extract_mode == "protein":
        raw_terms = prompt_required("Product name(s) (comma-separated): ")
    else:
        raw_terms = prompt_required("Gene name(s) (comma-separated): ")

    gene_terms = normalize_gene_terms(raw_terms)
    gene_terms_lower = set(g.lower() for g in gene_terms)

    taxid = prompt_optional_str("TaxID (optional, recommended): ")
    organism = prompt_optional_str("Organism name (recommended if no TaxID): ")

    max_assemblies = prompt_optional_int("Max RefSeq assemblies to scan: ") or ASSEMBLY_RETMAX
    max_sequences = prompt_optional_int("Max extracted sequences to output (optional): ")

    output_base = input("Output filename base [download]: ").strip() or "download"
    output_base = output_base.removesuffix(".fasta").removesuffix(".xlsx")
    out_fasta = Path(output_base + ".fasta")
    out_xlsx = Path(output_base + ".xlsx")

    make_code_version = (input("Create code-version FASTA? [y/N]: ").strip().lower() == "y")
    code_prefix = None
    out_code_fasta = None
    coded_records: list[SeqRecord] = []
    code_counter = 1
    if make_code_version:
        code_prefix = prompt_required("Code name (e.g., SAU): ").strip()
        out_code_fasta = Path(output_base + "_code.fasta")

    assembly_query = build_assembly_query(organism=organism, taxid=taxid)
    print("\nAssembly query (RefSeq-only):")
    print(assembly_query)
    print()

    try:
        assembly_ids = esearch_ids(db="assembly", term=assembly_query, retmax=max_assemblies)
    except Exception as exc:
        print(f"Error searching assemblies: {exc}")
        return

    if not assembly_ids:
        print("No RefSeq assemblies found for that organism/TaxID.")
        return

    extracted_records: list[SeqRecord] = []
    qc_rows: list[dict] = []

    assemblies_scanned = 0
    assemblies_with_hits: set[str] = set()
    genbank_records_scanned = 0
    cds_features_scanned = 0
    matches_found = 0
    skipped_no_translation = 0
    skipped_extract_failed = 0

    try:
        for asm_id in assembly_ids:
            assemblies_scanned += 1

            nuccore_ids = elink_assembly_to_nuccore_ids(asm_id)
            if not nuccore_ids:
                continue

            for start in range(0, len(nuccore_ids), GB_FETCH_BATCH):
                batch_ids = nuccore_ids[start : start + GB_FETCH_BATCH]
                gb_records = fetch_genbank_batch_records_with_retry(batch_ids)

                for gb_rec in gb_records:
                    genbank_records_scanned += 1

                    org_name = gb_rec.annotations.get("organism", "")
                    genome_acc = gb_rec.id
                    rec_desc = gb_rec.description

                    for feat in gb_rec.features:
                        if feat.type != "CDS":
                            continue
                        cds_features_scanned += 1

                        match, matched_by = feature_matches_gene(feat, gene_terms_lower)
                        if not match:
                            continue
                        matches_found += 1

                        gene_q = (feat.qualifiers.get("gene", [""])[0] or None)
                        locus_tag = (feat.qualifiers.get("locus_tag", [""])[0] or None)
                        product = (feat.qualifiers.get("product", [""])[0] or None)
                        protein_id = (feat.qualifiers.get("protein_id", [""])[0] or None)

                        strand = getattr(feat.location, "strand", None)
                        strand_s = "" if strand is None else str(strand)
                        loc_s = str(feat.location)

                        if extract_mode == "protein":
                            prot = extract_protein_from_feature(feat)
                            if not prot:
                                skipped_no_translation += 1
                                continue
                            seq = Seq(prot)
                        else:
                            cds_nt = extract_cds_nucleotide(gb_rec, feat)
                            if cds_nt is None:
                                skipped_extract_failed += 1
                                continue
                            seq = cds_nt

                        header = make_fasta_header(genome_acc, org_name, gene_q, product, protein_id, locus_tag)
                        extracted_records.append(SeqRecord(seq, id=genome_acc, description=header))
                        assemblies_with_hits.add(asm_id)

                        code_value = ""
                        if make_code_version and code_prefix:
                            code_value = f"{code_prefix}{code_counter}"
                            code_counter += 1
                            coded_records.append(SeqRecord(seq, id=code_value, description=""))

                        qc_rows.append(
                            {
                                "organism": org_name,
                                "genome_accession": genome_acc,
                                "code": code_value,
                                "record_description": rec_desc,
                                "matched_by": matched_by,
                                "gene": gene_q or "",
                                "locus_tag": locus_tag or "",
                                "product": product or "",
                                "protein_id": protein_id or "",
                                "cds_location": loc_s,
                                "strand": strand_s,
                                "length": len(seq),
                            }
                        )

                        if max_sequences and len(extracted_records) >= max_sequences:
                            raise StopIteration

            print(
                f"Scanned assemblies: {assemblies_scanned}/{len(assembly_ids)} | extracted: {len(extracted_records)}",
                end="\r",
            )

    except StopIteration:
        print()

    except Exception as exc:
        print(f"\nError during download/extraction: {exc}")
        return

    print("\n")
    print("=== Run summary (QC) ===")
    print(f"Assemblies scanned: {assemblies_scanned}")
    print(f"Assemblies with ≥1 hit: {len(assemblies_with_hits)}")
    print(f"GenBank records parsed: {genbank_records_scanned}")
    print(f"CDS features scanned: {cds_features_scanned}")
    print(f"CDS matches found: {matches_found}")
    print(f"Sequences extracted (written to FASTA): {len(extracted_records)}")
    if extract_mode == "protein":
        print(f"Skipped (no /translation): {skipped_no_translation}")
    if skipped_extract_failed:
        print(f"Skipped (CDS extract failed): {skipped_extract_failed}")
    print()

    if not extracted_records:
        print("No matching CDS found in scanned RefSeq assemblies.")
        return

    SeqIO.write(extracted_records, out_fasta, "fasta")
    if make_code_version and out_code_fasta:
        SeqIO.write(coded_records, out_code_fasta, "fasta")
    write_excel(qc_rows, out_xlsx)

    print(f"Done. Extracted {len(extracted_records)} sequences from {assemblies_scanned} RefSeq assemblies.")
    print(f"FASTA: {out_fasta.resolve()}")
    if make_code_version and out_code_fasta:
        print(f"Code FASTA: {out_code_fasta.resolve()}")
    print(f"Excel: {out_xlsx.resolve()}")


if __name__ == "__main__":
    main()