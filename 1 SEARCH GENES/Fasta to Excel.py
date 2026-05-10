# Script para pasar todos los datos de las secuencias FASTA de las blaZ_filtradas a un Excel

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

ruta_fasta = "blaZ_filtradas.fasta" 

secuencias =[]

accession = "" 
nombre = "" 
secuencia = ""

with open(ruta_fasta, "r") as archivo:
    for linea in archivo:
        linea = linea.strip() # Quita los espacios en blanco
        if linea.startswith(">"):
            if accession:
                secuencias.append({
                    "accession":accession,
                    "name": nombre,
                    "length": len(secuencia)
                })
            # Esto lo que hace es guardar dentro de cada variable vacía que tenemos todos los datos del header del FASTA
            # Hay que guardar por separado los accession numbers con el resto del texto, y lo único que los separa es un espacio " "
            # Por tanto tenemos que:

            partes = linea[1:].split(" ",1) # "Sáltate el primer caracter con [1:] (el ">"), y haz 1 solo split donde haya un " "
            accession = partes [0]
            nombre = partes [1]
            secuencia = "" # Con esto reiniciamos la secuencia y hacemos que vuelva a estar vacía

        else: # Todo lo que no empiece por ">" seran nucleotidos, por tanto los vamos sumando dentro de "secuencia"
            secuencia = secuencia + linea
    
    # Al  ser un bucle for, como después de la ultima secuencia no encuentra ningun ">" no la guarda
    # Ponemos de nuevo lo mismo FUERA para que la guarde después de que acabe

    if accession: 
        secuencias.append({
            "accession":accession,
            "name": nombre,
            "length": len(secuencia)   
        })



# Ahora, con toda esta informacion dentro de secuencias, hay que crear el archivo de excel

wb = openpyxl.Workbook()
hoja = wb.active
hoja.title = "Secuencias"

# Definir estilos para el encabezado
estilo_encabezado = Font(name="Arial", bold=True, color="FFFFFF", size=11)
fondo_encabezado = PatternFill("solid", start_color="2E4057")  # Azul oscuro
centrado = Alignment(horizontal="center", vertical="center")

# Escribir los encabezados de columna
encabezados = ["Accession Number", "Name", "Length"]
for col, titulo in enumerate(encabezados, start=1):
    celda = hoja.cell(row=1, column=col, value=titulo)
    celda.font = estilo_encabezado
    celda.fill = fondo_encabezado
    celda.alignment = centrado

# Escribir los datos de cada secuencia
fondo_par = PatternFill("solid", start_color="EAF2FB")   # Azul muy claro (filas pares)

for fila, seq in enumerate(secuencias, start=2):
    hoja.cell(row=fila, column=1, value=seq["accession"])
    hoja.cell(row=fila, column=2, value=seq["name"])
    hoja.cell(row=fila, column=3, value=seq["length"])

    # Aplicar fondo alternado para facilitar la lectura
    if fila % 2 == 0:
        for col in range(1, 4):
            hoja.cell(row=fila, column=col).fill = fondo_par

    # Centrar la columna de accession y length
    hoja.cell(row=fila, column=1).alignment = centrado
    hoja.cell(row=fila, column=3).alignment = centrado

    # Fuente Arial para los datos
    for col in range(1, 4):
        hoja.cell(row=fila, column=col).font = Font(name="Arial", size=10)

# Ajustar el ancho de las columnas automáticamente
hoja.column_dimensions["A"].width = 20
hoja.column_dimensions["B"].width = 90
hoja.column_dimensions["C"].width = 12

# Fijar la primera fila (encabezado) al hacer scroll
hoja.freeze_panes = "A2"

# Guardar el archivo Excel
ruta_salida = "blaZ_secuencias_2.0.xlsx"
wb.save(ruta_salida)


